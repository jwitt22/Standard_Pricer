from flask import Flask, render_template, request, send_from_directory, redirect, url_for, flash
import os
import pandas as pd
import zipfile
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import boto3

s3 = boto3.client('s3')

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

# Define the functions and lists
def is_building_unit(value):
    if pd.isna(value):
        return False
    return bool(re.match(r"(Building|Bld) \d+ Unit \w+", value))

def is_room_name(value):
    if pd.isna(value) or not isinstance(value, str):
        return False
    return not (value.startswith("Bld") or value.startswith("Building") or value.startswith("SPACE COUNT"))

def upload_to_s3(file_name, bucket, object_name=None):
    """Upload a file to an S3 bucket

    :param file_name: File to upload
    :param bucket: Bucket to upload to
    :param object_name: S3 object name. If not specified then file_name is used
    :return: True if file was uploaded, else False
    """

    # If S3 object_name was not specified, use file_name
    if object_name is None:
        object_name = file_name

    # Upload the file
    try:
        response = s3.upload_file(file_name, bucket, object_name)
    except Exception as e:
        print(e)
        return False
    return True


# Updated product list
products = [
    "Top Track 57\" - Silver",
    "Top Track 37\" - Silver",
    "Top Track Cover - Silver",
    "Standard 84\"",
    "Standard 72\"",
    "Standard 48\" - Silver",
    "Standard 36\" - Silver",
    "Standard Connector - Silver",
    "Wall Clip (Sold in Sets of 2)",
    "Closet Rod 2' - Silver",
    "Closet Rod 3' - Silver",
    "Closet Rod Holder, Pair - Silver",
    "Shelf 12\"x24\" - Silver",
    "Shelf 12\"x36\" -Silver",
    "Bracket 12\" - Silver",
    "Bracket Cover 12\", L/R - Silver",
    "Shelf Liner 12\"x24\"",
    "Shelf 16\"x24\" - Silver",
    "Shelf 16\"x36\" - Silver",
    "Bracket 16\" - Silver",
    "Bracket Cover 16\", L/R - Silver",
    "Shelf Liner 16\"x24\"",
    "Top Track Anchors",
    "Wall Clip Anchors, Pair - Silver",
    "Metal Drawer Frame 12\"x24\"",
    "Metal Mesh Basket 24\"",
    "Stationary Shoe Rack 24\" 2 Tier",
    "Stationary Shoe Rack 18\" 1 Tier",
    "Stationary Shoe Rack 24\" Single Row",
    "Gliding Shoe Rack 24\"",
    "Wood Fascia 18\"",
    "Wood Fascia 24\"",
    "Wood Fascia 36\"",
    "Wood Shelf 16\" x 24\"",
    "Wood Shelf 16\" x 30\"",
    "Solid Cubbie 7\"",
    "Solid Cubbie 7\" - 18\" wide",
    "Solid Cubbie 10\"",
    "Solid Cubbie 10\" - 18\" wide",
    "Wood Drawer 7\"",
    "Wood Drawer 7\" - 18\" wide",
    "Wood Drawer 10\"",
    "Wood Drawer 10\" - 18\" wide",
    "Drawer Frame 18\"",
    "Drawer Frame 24\""
]

# Updated prices list
prices = [
    "$21.00", "$15.00", "$1.00", "$27.00", "$22.00", "$15.00", "$11.00", "$3.00", "$4.50", "$12.00",
    "$16.00", "$7.00", "$18.00", "$27.00", "$5.00", "$3.00", "$2.00", "$22.00", "$34.00", "$7.00",
    "$4.00", "$3.00", "$1.50", "$1.00", "$42.00", "$33.00", "$44.00", "$18.00", "$26.00", "$45.00",
    "$36.00", "$39.00", "$49.00", "$78.00", "$97.00", "$118.00", "$105.00", "$131.00", "$120.00", "$189.00",
    "$159.00", "$218.00", "$182.00", "$36.00", "$42.00"
]

ignored_items = [
    "SPACE COUNT",
    "Wall D 10ft+",
    "Wall C 10ft+",
    "Wall A 10ft+",
    "Wall D 5-10ft",
    "Wall A 0-5ft",
    "Wall D 0-5ft",
    "Wall B 0-5ft",
    "Wall B 10ft+",
    "Wall A 5-10ft ",
    "Wall B 5-10ft",
    "Wall C 0-5ft",
    "Wall C 5-10ft"
]

def initialize_product_dict(products):
    return {product: 0 for product in products}

def extract_building_and_unit(value):
    match = re.search(r"(Building|Bld) (\d+) Unit (\w+)", value)
    if match:
        return match.group(2), match.group(3)  # Returns building number and unit letter/number
    return None, None

# Your existing functions like collect_data_flattened, save_data, etc. remain unchanged
def collect_data_flattened(df, products):
    data_collection = {}
    current_building = None
    current_unit = None
    room_data = []
    current_room_data = initialize_product_dict(products)
    room_wall_codes = []  # Initialize the room_wall_codes list here
    code = None  # Initialize the code variable
    wall_designation = ""  # Initialize the wall_designation variable
    
    for _, row in df.iterrows():
        # Detect a new unit
        
        if isinstance(row['Group'], str) and ("Bld" in row['Group'] or "Building" in row['Group']):
            # Save the previous unit's data if available
            print(f"Detected New Unit: {current_building} - {current_unit}")
            #print(f"Detected: {row['Group']}")
            if current_building and current_unit:
                data_collection[(current_building, current_unit)] = (room_data, room_wall_codes)  # Save both room data and wall codes
                room_data = []
                room_wall_codes = []  # Reset the room_wall_codes list for the next unit
            current_building, current_unit = extract_building_and_unit(row['Group'])
            wall_designation = ""  # Reset the wall_designation for the new unit
            continue

        # If the 'Assembly name' is not NaN, then it's a product row and we can extract the code
        if not pd.isna(row['Assembly name']):
            code = row['Assembly name']

        # Update the wall_designation if a new one is found
        if "Wall" in str(row['Item name']) and "Clip" not in str(row['Item name']):
            wall_designation = row['Item name']
        
        # Construct the full room-wall-code
        room_name = row['Group']
        #print(f"Row Group: {row['Group']}, Code: {code}, Wall Designation: {wall_designation}")
        full_room_wall_code = f"{room_name} - {code} - {wall_designation}"
        #print(f"Constructed Room-Wall-Code: {full_room_wall_code}")

        # Collect product data for the current room
        if row['Item name'] in current_room_data:
            current_room_data[row['Item name']] += row['QTY']

        # Detect a new room (by checking if the next row has a different room name)
        if _ == len(df) - 1 or row['Group'] != df.iloc[_ + 1]['Group']:
            print(f"Detected end of room: {row['Group']}")
            room_wall_codes.append(full_room_wall_code)
            room_data.append(current_room_data.copy())
            current_room_data = initialize_product_dict(products)

    # Save the last room's data
    if current_room_data:
        room_data.append(current_room_data)

    # Save the last unit's data
    if current_building and current_unit:
        data_collection[(current_building, current_unit)] = (room_data, room_wall_codes)
    
    #print(f"Collected data for {full_room_wall_code}: {current_room_data}")
    #print(f"Building-Unit pairs: {data_collection.keys()}")
    #print(data_collection)
    return data_collection

output_directory = PROCESSED_FOLDER

def save_data(data_collection, products, prices):
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')  # This will give you a format like '20231005123059'
    output_directory = "processed"
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    saved_files = []
    print(f"Number of buildings/units to process: {len(data_collection)}")
    processed_files = set()  # Use a set to track processed files

    for (building, unit), (room_data_list, room_wall_codes) in data_collection.items():
        file_path = os.path.join(output_directory, f"Building_{building}_{timestamp}.xlsx")
        # Load the workbook once per building
        if file_path not in processed_files:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
            else:
                wb = Workbook()
                wb.remove(wb.active)  # Remove the default sheet
            processed_files.add(file_path)
        else:
            wb = load_workbook(file_path)  # Load the workbook again for the next unit

        # Check if the sheet for the current unit exists, if not create one
        sheet_name = f"Unit {unit}"
        if sheet_name in wb.sheetnames:
            print(f"Warning: Duplicate data for Building {building}, Unit {unit}. Skipping this unit.")
            continue
        else:
            ws = wb.create_sheet(title=sheet_name)


        
        # Styling for the Unit header
        header_font = Font(bold=True, size=14)
        ws.append([f"Unit {unit}"])
        ws["A1"].font = header_font
        
        # Styling for the titles
        title_font = Font(size=12)
        title_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        ws.append(['Product', 'Quantity', 'List Price', 'Total'])
        for cell in ws["A2:D2"][0]:
            cell.font = title_font
            cell.fill = title_fill
        
        total_list_price = 0
        #print(f"Data for Building {building}, Unit {unit}: {room_data_list}")

        for room_index, item_quantities in enumerate(room_data_list):
            #print(f"Length of room_data_list: {len(room_data_list)}")
            #print(f"Length of room_wall_codes: {len(room_wall_codes)}")
            #print(f"Current room_index: {room_index}")


            try:
                room_wall_code = room_wall_codes[room_index]
                #print(f"Processing data for {room_wall_code}: {item_quantities}")
            except IndexError:
                print(f"Warning: No wall code found for Building {building}, Unit {unit}, Room index {room_index}. Skipping this room.")
                continue
            # ... [rest of the code for processing the room]

            ws.append([room_wall_code])
            for cell in ws[ws.max_row]:  # Style the last row (which is the room-wall-code)
                cell.font = Font(bold=True)

            for product in products:
                qty = item_quantities.get(product, 0)
                price = float(prices[products.index(product)].replace("$", ""))
                total = qty * price
                total_list_price += total
                ws.append([product, qty, f"${price:.2f}", f"${total:.2f}"])
            
            # Add an empty row between room data for clarity
            ws.append([])

        # Append the three rows at the bottom
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        
        summary_rows = [
            [None, None, "List Price>>>", f"${total_list_price:.2f}"],
            [None, None, "Discounted Price>>>", ""],
            [None, None, "Asking Price>>>", f"${total_list_price:.2f}"]
        ]
        
        for row_data in summary_rows:
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.fill = black_fill
                cell.font = white_font

        # Adjusting the width of column A
        max_length = max([len(product) for product in products])
        ws.column_dimensions['A'].width = max_length

        # Adjust the width of column C
        ws.column_dimensions['C'].width = len("Discounted Price>>>") + 2  # +2 for a bit of padding
        wb.save(file_path)
        # After saving the file with wb.save(file_path)
        upload_to_s3(file_path, 'quotefilecache')

        print(f"Saved file: {file_path}")
        saved_files.append(file_path)
    return saved_files

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # If the user does not select a file, the browser submits an empty file without a filename.
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = file.filename
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Process the file
            df = pd.read_excel(filepath)
            df['Group'] = df['Group'].ffill()
            df['Building'], df['Unit'] = zip(*df['Group'].apply(extract_building_and_unit))
            df['Building'] = df['Building'].fillna(method='ffill')
            df['Unit'] = df['Unit'].fillna(method='ffill')
            
            # Collect data from the flattened dataframe
            data_collection = collect_data_flattened(df, products)
            # Save the collected data to Excel
            saved_files = save_data(data_collection, products, prices)  # This will return a list of saved file paths

            # Create a ZIP archive containing all the saved files
            zip_filename = f"processed_files_{datetime.now().strftime('%Y%m%d%H%M%S')}.zip"
            zip_path = os.path.join(app.config['PROCESSED_FOLDER'], zip_filename)
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                for saved_file in saved_files:
                    zipf.write(saved_file, os.path.basename(saved_file))

            # Render a template that displays a link to download the ZIP archive
            return render_template('download_zip.html', zip_filename=zip_filename)

    return render_template('upload.html')


@app.route('/downloads/<filename>')
def download_file(filename):
    return send_from_directory(app.config['PROCESSED_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
